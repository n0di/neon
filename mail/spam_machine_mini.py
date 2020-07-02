import openpyxl
import smtplib
import random
import re
from email.mime.text import MIMEText
from email.header import Header

mail_g = 'amanxojayevn@gmail.com'
mail_pass = 'nodir19969796'

def auth_cl():
	mail_g = 'amanxojayevn@gmail.com'
	mail_pass = 'nodir19969796'


def list_cl():
	pass


def spam():
	print('Начинаем работу... \n')
	global recipient_list
	recipient_list = []
	with open('rec_list.txt', 'r') as f:
		for mail in f:
			mail = mail.replace('[', '').replace('\'', '').replace(']', '').replace('\n', '')
			recipient_list.append(mail)
			print(mail)
	path = 'sender_base.xlsx'  # Какой файл адресов для рассылки читаем?
	workbook = openpyxl.load_workbook(path)  # Собственно - читаем сам файл
	sheets_list = workbook.sheetnames  # Получаем список всех листов в книге
	
	global data_from_row, sheet, column_count, random_column, mail_adress_recipient, column_a, column_b, work_column_a, work_column_b, mail_server #Делаем глобальные переменные (уточнить)
	sheet = workbook[sheets_list[0]]  # Делаем активным самый первый лист в книге
	column_count = sheet.max_row
	print(column_count)
	random_column = 1 #random.randrange(2, column_count) #получаем случайную строку
	random_column = str(random_column)

	column_a = 'A'
	column_b = 'B'

	work_column_a = column_a + random_column
	#work_column_a = str(work_column_a)

	work_column_b = column_b + random_column
	#work_column_b = str(work_column_b)

	#Определяем почтовый сервер
	print(work_column_a)

	data_from_row = sheet[work_column_a].value
	regxp = '(@\w+.\w+)'
	mail_server = re.findall(regxp, data_from_row)
	print('Почтовый сервер:', mail_server)
	mail_server = str(mail_server)
	mail_server = mail_server.replace('[', '').replace(']', '').replace('\'','')
	print('Определили почтовый сервер:', mail_server)


	if mail_server == '@gmail.com':
		print('Работаем через Gmail')
		mailsender = smtplib.SMTP('smtp.gmail.com', 587)
		mailsender.starttls()
		mailsender.login(work_column_a, work_column_b)
		mail_subject = 'Тема сообщения'
		mail_body = 'Текст сообщения'
		msg = MIMEText(mail_body, 'plain', 'utf-8')
		msg['Subject'] = Header(mail_subject, 'utf-8')
		mailsender.sendmail(work_column_a, mail, msg.as_string())
		mailsender.quit()
		print('Сообщение на адрес', mail, 'отправлено')

	elif mail_server == '@yandex.ru':
		print('Работаем через Yandex')
		mailsender = smtplib.SMTP('smtp.yandex.ru', 587)
		mailsender.starttls()
		mailsender.login(work_column_a, work_column_b)
		mail_subject = 'Тема сообщения'
		mail_body = 'Текст сообщения'
		msg = MIMEText(mail_body, 'plain', 'utf-8')
		msg['Subject'] = Header(mail_subject, 'utf-8')
		mailsender.sendmail(work_column_a, mail, msg.as_string())
		mailsender.quit()
		print('Сообщение на адрес', mail, 'отправлено')

	elif mail_server == '@mail.ru':
		print('Работаем через Mail.ru')
		mailsender = smtplib.SMTP('smtp.mail.ru', 587)
		mailsender.starttls()
		mailsender.login(work_column_a, work_column_b)
		mail_subject = 'Тема сообщения'
		mail_body = 'Текст сообщения'
		msg = MIMEText(mail_body, 'plain', 'utf-8')
		msg['Subject'] = Header(mail_subject, 'utf-8')
		mailsender.sendmail(work_column_a, mail, msg.as_string())
		mailsender.quit()
		print('Сообщение на адрес', mail, 'отправлено')

	elif mail_server == '@outlook.com':
		print('Работаем с Outlook.com')
		mailsender = smtplib.SMTP('smtp.outlook.com', 587)
		mailsender.starttls()
		mailsender.login(work_column_a, work_column_b)
		mail_subject = 'Тема сообщения'
		mail_body = 'Текст сообщения'
		msg = MIMEText(mail_body, 'plain', 'utf-8')
		msg['Subject'] = Header(mail_subject, 'utf-8')
		mailsender.sendmail(work_column_a, mail, msg.as_string())
		mailsender.quit()
		print('Сообщение на адрес', mail, 'отправлено')
def work():
	spam()

work()
