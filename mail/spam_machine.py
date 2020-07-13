import openpyxl
import random
import re
import smtplib
import os
from email.mime.text import MIMEText
from email.header import Header

max_mail = int(input('how many adress you need?: '))

def create_sender_list():
	start_line = 1
	wb = openpyxl.Workbook()
	column_a = 'A'
	column_b = 'B'
	while start_line <= max_mail:
		sheets_list = wb.sheetnames  # Get a list of all sheets in a file
		sheet_active = wb[sheets_list[0]]  # Getting started with the very first
		mail_server_list = ['gmail.com', 'outlook.com']
		random_value = random.randrange(1, 10)
		random_mail = random.sample('abcdefghijklmnopqrstuvwxyz0123456789', random_value)
		random_mail = ''.join(random_mail)
		random_mail_server = mail_server_list[random.randrange(0, len(mail_server_list))]
		random_mail = random_mail + '@' + random_mail_server
		start_line = start_line + 1
		start_line = str(start_line)
		sheet_active[column_a + start_line] = random_mail
		start_line = int(start_line)

		#generate pass for mail
		random_value = random.randrange(8, 12)
		password_for_mail = random.sample('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789!@#$%^&*', random_value)
		password_for_mail = ''.join(password_for_mail)

		#And now we write everything to a file
		start_line = str(start_line)
		sheet_active[column_b + start_line] = password_for_mail
		start_line = int(start_line)
		print(random_mail, ':', password_for_mail, ' - ready')

		wb.save('sender_base.xlsx')
	print('Database of mailing addresses for sending created.\n')

def create_recipient_list():
	start_line = 1
	while start_line <=  max_mail * 3:
		random_value = random.randrange(1, 10)
		random_mail = random.sample('abcdefghijklmnopqrstuvwxyz0123456789', random_value)
		random_mail = ''.join(random_mail)
		random_mail = random_mail + '@gmail.com'
		with open('recipient_list.txt', 'a', encoding='utf8') as f:
			f.write(random_mail + '\n')
		start_line = start_line + 1
	print('Recipient database created :)')

def spam():
	print('Getting started... \n')
	global recipient_list

	recipient_list = []
	with open('recipient_list.txt', 'r', encoding='utf8') as f:
		for mail in f:
			mail = mail.replace('[', '').replace('\'', '').replace(']', '').replace('\n', '')

			recipient_list.append(mail)
			print(mail)

			#Getting started with the address file to send
			path = 'sender_base.xlsx'  # What mailing list file are we reading?
			workbook = openpyxl.load_workbook(path)  # Actually - we read the file itself
			sheets_list = workbook.sheetnames  # Get a list of all the sheets in the book.
			global data_from_row, sheet, column_count, random_column, mail_adress_recipient, column_a, column_b, work_column_a, work_column_b, mail_server #Делаем глобальные переменные (уточнить)
			sheet = workbook[sheets_list[0]]  # Making the very first sheet in the book active
			column_count = sheet.max_row
			print(column_count)
			random_column = random.randrange(2, column_count) # we get a random string
			random_column = str(random_column)

			column_a = 'A'
			column_b = 'B'

			work_column_a = column_a + random_column
			#work_column_a = str(work_column_a)

			work_column_b = column_b + random_column
			#work_column_b = str(work_column_b)

			# Defining a mail server
			print(work_column_a)

			data_from_row = sheet[work_column_a].value

			regxp = '(@\w+.\w+)'
			mail_server = re.findall(regxp, data_from_row)
			print(f'Mail server: {mail_server}')
			mail_server = str(mail_server)
			mail_server = mail_server.replace('[', '').replace(']', '').replace('\'','')
			print(f'Defined a mail server: {mail_server}')

			#And now we start sending out with a random line taking into account the mail server:
			if mail_server == '@gmail.com':
				print('We work through Gmail')
				mailsender = smtplib.SMTP('smtp.gmail.com', 587)
				mailsender.starttls()
				mailsender.login(work_column_a, work_column_b)
				mail_subject = 'Message subject'
				mail_body = 'Text message'
				msg = MIMEText(mail_body, 'plain', 'utf-8')
				msg['Subject'] = Header(mail_subject, 'utf-8')
				mailsender.sendmail(work_column_a, mail, msg.as_string())
				mailsender.quit()
				print(f'Message sent to {mail}')

			elif mail_server == '@outlook.com':
				print('We work through Outlook.com')
				mailsender = smtplib.SMTP('smtp.outlook.com', 587)
				mailsender.starttls()
				mailsender.login(work_column_a, work_column_b)
				mail_subject = 'Message subject'
				mail_body = 'Text message'
				msg = MIMEText(mail_body, 'plain', 'utf-8')
				msg['Subject'] = Header(mail_subject, 'utf-8')
				mailsender.sendmail(work_column_a, mail, msg.as_string())
				mailsender.quit()
				print(f'Message sent to {mail}')

def work():
	create_sender_list() # Create a list of addresses from which we will send
	#create_recipient_list() # Create a list of addresses to which we will send
	spam() #start

work()
